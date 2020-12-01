from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fills import PatternFill

wb = Workbook()
outputFile: Worksheet = wb.active
inputFile = load_workbook('INPUT.xlsx')

sheetInput = inputFile.get_sheet_by_name('Sheet1')

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

orangeFill = PatternFill(start_color='ff8000',
                   end_color='ff8000',
                   fill_type='solid')

yellowFill = PatternFill(start_color='ffff00',
                   end_color='ffff00',
                   fill_type='solid')


keys = []
yellow = 0
orange = 0
red = 0
for i in range(0, 5):
    col = i + 2
    keys.append(outputFile.cell(row=3, column=col))

dict1 = {keys[0]: 'String 1',
         keys[1]: 'String 2',
         keys[2]: 'String 3',
         keys[3]: 'String 4',
         keys[4]: 'String 5'
         }

for j in range(2, 106):
    cell = sheetInput.cell(row=j, column=16).value
    data = str(cell)
    split = data.split(';')

    skip = []
    for k in split:
        currentColumn = 1
        result = False
        for key in dict1:
            if k == dict1[key]:
                result = True
                outputFile.cell(row=j, column=currentColumn).value = 1
                skip.append(currentColumn)
            else:
                if currentColumn not in skip:
                    outputFile.cell(row=j, column=currentColumn).value = 0
            currentColumn+=1
    if len(skip) == 4:
        outputFile.cell(row=j, column=1).fill = yellowFill
        yellow += 1
    elif len(skip) == 5:
            outputFile.cell(row=j, column=1).fill = orangeFill
            orange += 1
    elif len(skip) > 5:
            outputFile.cell(row=j, column=1).fill = redFill
            red += 1

wb.save('testFile.xlsx')

print('yellow: ' + str(yellow))
print('orange: ' + str(orange))
print('red: ' + str(red))

"""
Code for alternative type of data/columns from input file

keys = [0, 1, 2, 3]
yellow = 0
orange = 0
red = 0


dict1 = {1: 'String 1',
         2: 'String 2'
         } 

for j in range(2, 106):
    cell = sheetInput.cell(row=j, column=7  ).value
    data = str(cell)
    split = data.split(';')

    skip = []
    for k in split:
        currentColumn = 1
        for key in dict1:
            if k == dict1[1]:
                outputFile.cell(row=j, column=currentColumn).value = 1
            else:
                outputFile.cell(row=j, column=currentColumn).value = 2
           
"""